#%%
import openpyxl
import glob
import shutil
from natsort import natsorted
import re
sheet1_row = 1
sheet1_column = 1
after_row = 1
after_column = 1
column_list = ['A','B','C','D']
files = glob.glob('F:/test/*.xlsx')
new_file = openpyxl.Workbook()

for name in natsorted(files):#ファイルを上から順番に取得
    before = openpyxl.load_workbook(name)
    sheet1 = before['書面']
    for row in range(2,48):
        syurui = sheet1.cell(row,2).value
        if syurui == 'ｿｹｯﾄｽｸﾘｭｰ':
                No = str(sheet1.cell(10,1).value)
                s = sheet1.cell(row,4).value#サイズ取得
                quanity = sheet1.cell(row,5).value#数量取得
                lists =s.split() #空白判定
                result1 = re.sub(r"\D", "", lists[0]) #空白の中の文字列1
                result2= re.sub(r"\D", "", lists[1]) #空白の後の文字列
                syurui = result1 +"x"+result2
                print(syurui)
                print(quanity)
                after = new_file.active #新しく作ったファイルを利用可能に
                for row_num in range(1,12):#11列までセルサイズ変更
                    after.row_dimensions[row_num].height = 50
                after.cell(after_row,after_column).value=str(syurui)+'\n'+str(quanity)
                after.cell(after_row,after_column).alignment = openpyxl.styles.Alignment(wrapText=True)
                after_row +=1
                if after_row %11 ==0:
                    after.cell(after_row,after_column).value=str(syurui)+'\n'+str(quanity)
                    after.cell(after_row,after_column).alignment = openpyxl.styles.Alignment(wrapText=True)
                    after_column +=1
                    after_row = 1
                fileneme = "F:/test/" + No + ".xlsx"
                new_file.save(fileneme)
        else:
            pass
    after.page_margins.left = 0
    after.page_margins.right = 0
    after.page_margins.top = 0
    after.page_margins.bottom = 0

    after.page_margins.header = 0
    after.page_margins.footer = 0

    after.print_options.horizontalCentered = True
    after.print_options.verticalCentered = True
    new_file.save('F:/test/testafter.xlsx')
    shutil.move(name,'F:/test/testbefore')
# %%

file_name  = 1
for i in range(1,8):
    file_names = file_name +i
    print(file_names)
# %%
